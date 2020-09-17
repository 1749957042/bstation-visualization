#将获取的字典中的视频播放数据写入excel文档
import obtain   #引入爬取视频信息的文件
import pandas as pd
import openpyxl    #导出到Excel的关键库

# 运行此程序将会把指定区间内的有效视频信息存储输出到Excel文件中
home = 1  # 视频起始av号
end = 100    # 视频结束av号

#调用爬虫函数
data = obtain.main(home,end)

# 新建Excel工作薄
wb = openpyxl.Workbook()
# 设置表名
sheet = wb.create_sheet('B站视频信息爬取',0)
# 设置Excel列名列表
title_ls = ['视频标题','作者','av号','bv号','视频类型','播放量','弹幕数','评论数','收藏数','投币数','转发量','点赞数']
    
# 循环写入标题
for i in range(len(title_ls)):
        sheet.cell(1,1+i).value = title_ls[i]

# 循环将数据写入Excel文件
for row in range(len(data)):
        sheet.cell(row+2,1).value = data[row]['title'] #视频标题
        sheet.cell(row+2,2).value = data[row]['name']   #作者名字
        sheet.cell(row+2,3).value = data[row]['av']     #av号
        sheet.cell(row+2,4).value = data[row]['bv']     #bv号
        sheet.cell(row+2,5).value = data[row]['type']   #视频类型
        sheet.cell(row+2,6).value = data[row]['view']   #播放量
        sheet.cell(row+2,7).value = data[row]['danmu']  #弹幕数量
        sheet.cell(row+2,8).value = data[row]['comment']        #评论数量
        sheet.cell(row+2,9).value = data[row]['collect']        #收藏数量
        sheet.cell(row+2,10).value = data[row]['coin']  #投币数
        sheet.cell(row+2,11).value = data[row]['share'] #转发量
        sheet.cell(row+2,12).value = data[row]['like']  #点赞数

# 将数据保存到指定路径的文件中
wb.save('./bilibili_data'+'_'+str(home)+'_'+str(end)+'.xlsx')
wb.close()
